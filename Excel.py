import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string, get_column_letter


class Excel:
    def init(self) -> None:
        self.workbook = openpyxl.Workbook()

        self.featured_sheet = self.workbook.active
        self.featured_sheet.title = "Featured Stock"

        self.project_name = "NewSheet"


        self.featured_stock_creation()


    def featured_stock_creation(self):
        self.dropdown_creation(
            self.featured_sheet, 
            [
                "Exec Comp Aligned with ROIC", 
                "Safest Dividend Yield", 
                "Dividend Growth", 
                "Most Attractive/Most Dangerous"
            ], 
            "B2"
        )
        
        self.featured_sheet["B3"] = "Year Fraction"
        self.dropdown_creation(
            self.featured_sheet, 
            ["0.25", "0.5", "0.75"], 
            "C3"
        )

        self.featured_sheet["B6"] = "Invested Capital Turns"


        self.save_project(f"{self.project_name}")


    def stock_file_manipulation(self):
        workbook = openpyxl.load_workbook('analysis/AZO.xlsx')
        sheet = workbook.active


        # TODO: remove 'self.correction' when the current column at the sheet can be used
        if round(sheet[(sheet["AU3"].value)[1:]].value) == round(sheet[(sheet["AV3"].value)[1:]].value):
            self.column = "AU"
            self.correction = 0

        else:
            self.column = "AV"
            self.correction = 1

        
        rows = [
            "Revenue 1Y", 
            "Revenue 3Y", 
            "Revenue 5Y", 
            "Revenue 10Y", 
            "Revenue 20Y", 
            "NOPAT 1Y", 
            "NOPAT 3Y", 
            "NOPAT 5Y", 
            "NOPAT 10Y", 
            "NOPAT 20Y", 
            "NMARGIN3Y", 
            "NMARGIN5Y", 
            "NMARGIN10Y"
        ]
        cell = 21

        for row in rows:
            sheet[f"{get_column_letter(column_index_from_string(f'{self.column}') - 1)}{cell}"] = row
            
            cell += 1

       

        sheet[f"{self.column}21"].value = f"={self.column}3/{get_column_letter(column_index_from_string(f'{self.column}') - (1 + self.correction))}3-1"
        sheet[f"{self.column}21"].number_format = '0%' 

        sheet[f"{self.column}22"].value = f"=({self.column}3/{get_column_letter(column_index_from_string(f'{self.column}') - (3 + self.correction))}3)^(1/3)-1"
        sheet[f"{self.column}22"].number_format = '0%'

        sheet[f"{self.column}23"].value = f"=({self.column}3/{get_column_letter(column_index_from_string(f'{self.column}') - (5 + self.correction))}3)^(1/5)-1"
        sheet[f"{self.column}23"].number_format = '0%'

        sheet[f"{self.column}24"].value = f"=({self.column}3/{get_column_letter(column_index_from_string(f'{self.column}') - (10 + self.correction))}3)^(1/10)-1"
        sheet[f"{self.column}24"].number_format = '0%'

        sheet[f"{self.column}25"].value = f"=({self.column}3/{get_column_letter(column_index_from_string(f'{self.column}') - (20 + self.correction))}3)^(1/20)-1"
        sheet[f"{self.column}25"].number_format = '0%'


        sheet[f"{self.column}26"].value = f"={self.column}7/{get_column_letter(column_index_from_string(f'{self.column}') - (1 + self.correction))}7-1"
        sheet[f"{self.column}26"].number_format = '0%'

        sheet[f"{self.column}27"].value = f"=({self.column}7/{get_column_letter(column_index_from_string(f'{self.column}') - (3 + self.correction))}7)^(1/3)-1"
        sheet[f"{self.column}27"].number_format = '0%'

        sheet[f"{self.column}28"].value = f"=({self.column}7/{get_column_letter(column_index_from_string(f'{self.column}') - (5 + self.correction))}7)^(1/5)-1"
        sheet[f"{self.column}28"].number_format = '0%'

        sheet[f"{self.column}29"].value = f"=({self.column}7/{get_column_letter(column_index_from_string(f'{self.column}') - (10 + self.correction))}7)^(1/10)-1"
        sheet[f"{self.column}29"].number_format = '0%'

        sheet[f"{self.column}30"].value = f"=({self.column}7/{get_column_letter(column_index_from_string(f'{self.column}') - (20 + self.correction))}7)^(1/20)-1"
        sheet[f"{self.column}30"].number_format = '0%'


        sheet[f"{self.column}31"].value = f"=average({get_column_letter(column_index_from_string(f'{self.column}') - (3 - self.correction))}8:{self.column}8)"
        sheet[f"{self.column}31"].number_format = '0%'

        sheet[f"{self.column}32"].value = f"=average({get_column_letter(column_index_from_string(f'{self.column}') - (5 - self.correction))}8:{self.column}8)"
        sheet[f"{self.column}32"].number_format = '0%'

        sheet[f"{self.column}33"].value = f"=average({get_column_letter(column_index_from_string(f'{self.column}') - (10 - self.correction))}8:{self.column}8)"
        sheet[f"{self.column}33"].number_format = '0%'


        workbook.save('analysis/AZO.xlsx')


    def stock_file_data(self):
        sheet_name = input("Write the stock sheet name \n ->")
        workbook = openpyxl.load_workbook(f'analysis/{sheet_name}.xlsx')
        sheet = workbook.active

        data = {}


        data['revenue_5y'] = sheet[f"{self.column}23"].value
        data['revenue_5y'] = self.calculate_cells(data['revenue_5y'], sheet, "div")

        data['nopat_5y'] = sheet[f"{self.column}28"].value
        data['nopat_5y'] = self.calculate_cells(data['nopat_5y'], sheet, "div")

        data['5_years_back'] = sheet[(sheet[f"{get_column_letter(column_index_from_string(f'{self.column}') - (5 + self.correction))}1"].value)[1:]].value

        data['current_percentage'] = round(float((sheet[(sheet[f"{get_column_letter(column_index_from_string(f'{self.column}')- 1)}8"].value)[1:]].value)) * 100)

        nmargin_list = sheet[f"{get_column_letter(column_index_from_string(f'{self.column}') - (6 + self.correction))}8:{get_column_letter(column_index_from_string(f'{self.column}')- 1)}8"]
        data['smaller_number'] = self.calculate_cells(nmargin_list, sheet, "smaller")

        data['smaller_number_year'] = self.calculate_cells(nmargin_list, sheet, "smaller_y", data['smaller_number'])


        columns = sheet["AO1:AV1"]

        for col in columns[0]:
            if sheet[(col.value)[1:]].value == data['smaller_number_year']:
                column = col


        data['roic_first'] = round(float(sheet[(sheet[f"{get_column_letter(column.column)}12"].value)[1:]].value * 100))

        data['roic_last'] = round(float(sheet[(sheet[f"{get_column_letter(column_index_from_string(f'{self.column}') - 1)}12"].value)[1:]].value * 100))


        return data


    def feature_stock_data(self, data):
        sheet_name = input("Write the new sheet name \n ->")
        workbook = openpyxl.load_workbook(f'created/{sheet_name}.xlsx')
        sheet = workbook.active


        year_list = sheet["D6:J6"]


        for year in year_list[0]:
            formatted_year = int(year.value)

            if formatted_year == data['smaller_number_year']:
                fist_column = get_column_letter(year.column)
                break


        data['first_investment'] = sheet[f"{fist_column}7"].value
        data['last_investment'] = sheet["J7"].value

        
        return data


    def calculate_cells(self, formula, sheet, type, add=None):
        if type == "div":
            formula = formula[1:].split(")")


            div1 = (formula[0][1:]).split("/")
            div1 = round(sheet[(sheet[div1[0]].value)[1:]].value) / round(sheet[(sheet[div1[1]].value)[1:]].value)

            div2 = (formula[1][2:]).split("/")
            div2 = int(div2[0]) / int(div2[1])

            result = round(((div1) ** (div2) - 1) * 100)

        elif type == "smaller":
            formatted_list = []

            for num in formula[0]:
                num = (sheet[num.value[1:]].value) * 100
                num = "{:.2f}".format(num)
                formatted_list.append(num)

            result = min(formatted_list)

        elif type == "smaller_y":
            for num in formula[0]:
                value = (sheet[num.value[1:]].value) * 100
                value = "{:.2f}".format(value)

                if value == add:
                    result = sheet[(sheet[f"{get_column_letter(num.column)}1"].value)[1:]].value
                    break


        return result


    def reading_company_file(self):
        workbook = openpyxl.load_workbook('analysis/AZO.xlsx')
        sheet = workbook.active


        return sheet["A2"].value


    def dropdown_creation(self, sheet, options, cell_range):
        data_validation = DataValidation(type='list', formula1='"'+','.join(options)+'"')

        sheet.add_data_validation(data_validation)

        data_validation.add(cell_range)


    def save_project(self, name):
        self.workbook.save(f'created/{name}.xlsx')
